import os
import smtplib
import ssl
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import ast
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import pandas as pd
from pathlib import Path

home = str(Path.home())
home = home.replace(os.sep, '/')
print(home)

ExcelFileName = "UserData"
loc = (ExcelFileName + '.xlsx')
wb = openpyxl.load_workbook(loc)

Sheetname="Data"
sheet = wb[Sheetname]
User_Name_Sheet={}
User_Name_Email={}
for x in range(2, 200):
    if sheet.cell(x,1).value == None:
        break
    else:
        User_Name_Sheet[sheet.cell(x, 1).value]=sheet.cell(x, 2).value
        User_Name_Email[sheet.cell(x, 1).value]=sheet.cell(x, 3).value

#print(User_Name_Sheet)
UserKeys=list(User_Name_Sheet.keys())
#print(UserKeys)
for user in range(0,len(UserKeys)):
    try:
        #-------------------To read content to send in e-Mail--------------------
        # Connecting with Main Report Data File
        ExcelFileName = "ReportData/"+User_Name_Sheet[UserKeys[user]]
        locx = (ExcelFileName + '.xlsx')
        wbx = openpyxl.load_workbook(locx)

        # Reading GlobalData tab of Main Report Data File
        date_str = pd.Timestamp.today().strftime('%m-%d-%Y')
        date_strPDF = pd.Timestamp.today().strftime("%d-%B-%Y-%I-%M-%p")
        Sheetname = "GlobalData"
        sheetx = wbx[Sheetname]
        for ix in range(1, 200):
            if sheetx.cell(ix, 1).value == None:
                break
            else:
                if sheetx.cell(ix, 1).value == "Report_Name":
                    print("Report_Name is: " + sheetx.cell(ix, 2).value)
                    FileDelete=sheetx.cell(ix, 2).value
                    Report_Name = sheetx.cell(ix, 2).value+'_'+date_str+'.pdf'
                    Report_Name1 = sheetx.cell(ix, 2).value + '_' + date_str + '.pdf'
                if sheetx.cell(ix, 1).value == "Email_From":
                    print("Email_From is: " + sheetx.cell(ix, 2).value)
                    Email_From = sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "Email_To":
                    print("Email_To is: " + sheetx.cell(ix, 2).value)
                    Email_To = sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "Email_Subject":
                    print("Email_Subject is: " + sheetx.cell(ix, 2).value)
                    Email_Subject = sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "Email_Content":
                    print("Email_Content is: " + sheetx.cell(ix, 2).value)
                    Email_Content = sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "GoogleAppCode":
                    print("GoogleAppCode is: " + sheetx.cell(ix, 2).value)
                    GoogleAppCode = sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "GoogleDriveFolderID":
                    print("GoogleDriveFolderID is: " + sheetx.cell(ix, 2).value)
                    GoogleDriveFolderID = sheetx.cell(ix, 2).value

        html = '''
            <html>
                <body>
                    <p>Hi Team</p 
                    <p>'''+Email_Content+'''<br /></p>
                    <p>To access old reports go to drive link given below <br /></p>
                    <p>'''+"https://drive.google.com/drive/folders/"+GoogleDriveFolderID+'''<br /><br /></p>
                    <p>Many Thanks <br/>'''+UserKeys[user]+'''</p>
                </body>
            </html>
            '''

        Report_Name=home+'/.jenkins/workspace/CreateReport/'+UserKeys[user]+"_"+Report_Name
        def attach_file_to_email(msg, attach,Report_Name, extra_headers=None):
            with open(attach, "rb") as f:
                file_attachment = MIMEApplication(f.read())
            file_attachment.add_header(
                "Content-Disposition",
                f"attachment; filename= {Report_Name}",
            )
            if extra_headers is not None:
                for name, value in extra_headers.items():
                    file_attachment.add_header(name, value)
            msg.attach(file_attachment)

        email_from = 'Test Automation Team'
        y = Email_To
        email_to = ast.literal_eval(y)
        SenderEmail=Email_From
        date_str = pd.Timestamp.today().strftime('%m-%d-%Y')
        msg = MIMEMultipart()
        msg['Subject']=Email_Subject+" "+date_str
        msg['From'] = email_from
        msg['To'] = ','.join(email_to)
        msg.attach(MIMEText(html, "html"))
        try:
            attach_file_to_email(msg, Report_Name,Report_Name1)
        except Exception as em:
            FileLoc=UserKeys[user] + "_" + Report_Name1
            attach_file_to_email(msg,FileLoc ,Report_Name1)

        #-----------------------------------------------------------------------

        # ------------------------To attach all in e-Mail-----------------------
        email_string = msg.as_string()
        context = ssl.create_default_context()
        # -----------------------------------------------------------------------

        # ----------------------------SMTP setup--------------------------------
        server=smtplib.SMTP_SSL('smtp.gmail.com',465)
        RandmStr=GoogleAppCode
        server.login(SenderEmail,RandmStr)
        #server.sendmail(email_from, email_to, email_string)
        print("Test Report sent")
        server.quit()

        #--------------------------GDrive setup-----------------------------------
        gauth = GoogleAuth()
        gauth.LoadCredentialsFile("mycreds.txt")
        if gauth.credentials is None:
            gauth.LocalWebserverAuth()
        elif gauth.access_token_expired:
            gauth.Refresh()
        else:
            gauth.Authorize()
        gauth.SaveCredentialsFile("mycreds.txt")
        drive = GoogleDrive(gauth)
        #--------------------------GDrive upload-----------------------------------
        upload_file_list = [Report_Name]
        try:
            for upload_file in upload_file_list:
                gfile = drive.CreateFile({'title': date_strPDF,'parents': [{'id': GoogleDriveFolderID}]})
                gfile.SetContentFile(upload_file)
                gfile.Upload()
        except:
            upload_file_list = [FileLoc]
            for upload_file in upload_file_list:
                gfile = drive.CreateFile({'title': date_strPDF,'parents': [{'id': GoogleDriveFolderID}]})
                gfile.SetContentFile(upload_file)
                gfile.Upload()
        #--------------------------------------------------------------------------

        sheetx.cell(row=1, column=5).value = date_str
        wbx.save(locx)

    except Exception as aaa:
        print(aaa)
        print(str(user) +" Report File not found for "+UserKeys[user])
