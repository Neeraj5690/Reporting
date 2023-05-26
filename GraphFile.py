from builtins import print

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib

Email_Content_NoFile="Good Evening !!!  You have missed to share report file for this week."
FileLink_NoFile="abc"
email_from_NoFile = 'Test Automation Team'
Email_From_NoFile="neeraj1wayitsol@gmail.com"
Email_Subject_NoFile="Weekly Report Not Sent"
GoogleAppCode_NoFile="tsiajyfnhywxctwi"

ExcelFileName = "UserData"
loc = (ExcelFileName + '.xlsx')
wb = openpyxl.load_workbook(loc)

Sheetname="Data"
sheet = wb[Sheetname]
User_Name={}
User_Name_Email={}
for x in range(2, 200):
    if sheet.cell(x,1).value == None:
        break
    else:
        print("User_Name : "+sheet.cell(x, 1).value)
        print("User_File : "+sheet.cell(x, 2).value)
        User_Name[sheet.cell(x, 1).value]=sheet.cell(x, 2).value
        User_Name_Email[sheet.cell(x, 1).value]=sheet.cell(x, 3).value

print(User_Name)
UserKeys=list(User_Name.keys())
print(UserKeys)
ModuleName=None
for user in range(0,len(UserKeys)):
    Project_Name = "QA"
    Graph_show = "No"
    BarGraph_color = "blue"

    print("---------------------------------------------------------")
    print("ModuleName " + str(ModuleName))

    try:
        ExcelFileName = "ReportData/"+User_Name[UserKeys[user]]
        locx = (ExcelFileName + '.xlsx')
        wbx = openpyxl.load_workbook(locx)

        Sheetname="GlobalData"
        sheetx = wbx[Sheetname]

        for ix in range(1, 200):
            if sheetx.cell(ix, 1).value == None:
                break
            else:
                if sheetx.cell(ix, 1).value == "Project_Name":
                    print("Project Name is: "+sheetx.cell(ix, 2).value)
                    Project_Name=sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "BarGraph_color":
                    BarGraph_color=sheetx.cell(ix, 2).value
                    print("BarGraph_color is "+BarGraph_color)
                if sheetx.cell(ix, 1).value == "Graph_show":
                    Graph_show=sheetx.cell(ix, 2).value
                    print("Graph_show is "+Graph_show)
                if sheetx.cell(ix, 1).value == "Graph_Type":
                    Graph_Type=sheetx.cell(ix, 2).value
                    print("Graph_Type is "+Graph_Type)

        Column_Name = []
        ModuleName = []
        Bugs_Count = {}
        Bugs_CountList = []
        data = {}

        if Graph_show == "Yes":
            Sheetname="ModulesData"
            sheetx = wbx[Sheetname]
            for ix1 in range(2, 200):
                if sheetx.cell(ix1, 1).value == None:
                    break
                else:
                    ModuleName.append(sheetx.cell(ix1, 1).value)
            print("ModuleName "+str(ModuleName))

            for ix2 in range(1, 20):
                if sheetx.cell(1,ix2).value == None:
                    break
                else:
                    Column_Name.append(sheetx.cell(1,ix2).value)
            print("Column_Name "+str(Column_Name))


            ColumnNumber=Column_Name.index("Bugs_Count")+1
            print("ColumnNumber: "+str(ColumnNumber))
            ifNothing=0
            for ix3 in range(1, len(ModuleName)+1):
                if sheetx.cell(ix3+1, 1).value == ModuleName[ix3-1]:
                    if sheetx.cell(ix3+1, ColumnNumber).value==None:
                        Bugs_Count[ModuleName[ix3 - 1]] = 0
                        Bugs_CountList.append(0)
                    else:
                        Bugs_Count[ModuleName[ix3-1]]=sheetx.cell(ix3+1, ColumnNumber).value
                        Bugs_CountList.append(sheetx.cell(ix3+1, ColumnNumber).value)
            print("Bugs_Count "+str(Bugs_Count))

            # Creating Module Vs Bugs Count Bar Graph
            print("ModuleName "+str(ModuleName))
            print("Bugs_CountList "+str(Bugs_CountList))
            Bugs_CountListSum=sum(Bugs_CountList)
            print(Bugs_CountListSum)
            print(data)

            if Graph_Type == "BarGraph":
                data = {'modules': ModuleName,
                    'bugs': Bugs_CountList
                    }
                print(data)
                df = pd.DataFrame(data)
                colors = [BarGraph_color]
                print(df)
                try:
                    plt.bar( df['modules'],df['bugs'], color=colors)
                except:
                    print("Color is invalid")
                    colors = ['blue']
                    plt.bar(df['modules'], df['bugs'], color=colors)

                ax = plt.gca()
                plt.draw()
                ax.set_xticklabels(df['modules'] , rotation = 55,fontsize=8)

                plt.title('Modules Vs Bugs Count', fontsize=10)
                #plt.xlabel('Modules', fontsize=8)
                plt.ylabel('Bugs', fontsize=8)
                plt.grid(False)
                plt.gcf().set_size_inches(5, 7)
                if Bugs_CountListSum == 0:
                    print("Bugs_CountListSumCount is 0")
                else:
                    plt.savefig(UserKeys[user]+'_ModuleVsBugsCount.jpg', dpi=500)
            elif Graph_Type == "PieChart":
                import numpy as np
                import matplotlib.pyplot as plt

                fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))
                Bugs_CountList1=Bugs_CountList
                for ix2 in range(0, len(Bugs_CountList)):
                    if Bugs_CountList[ix2] == 0:
                        print(ix2)
                        try:
                             ModuleName[ix2]=0
                        except:
                            pass
                Bugs_CountList = [elem for elem in Bugs_CountList if elem != 0]
                ModuleName = [elem for elem in ModuleName if elem != 0]
                print(ModuleName)
                print(Bugs_CountList)

                data = Bugs_CountList
                def func(pct, allvals):
                    absolute = int(pct / 100. * np.sum(allvals))
                    return "{:.1f}%\n({:d})".format(pct, absolute)
                wedges, texts, autotexts = ax.pie(data, autopct=lambda pct: func(pct, data),
                                                  textprops=dict(color="w"))
                ax.legend(wedges, ModuleName,
                          title="Modules",
                          loc="center left",
                          bbox_to_anchor=(1, 0, 0.5, 1))
                plt.setp(autotexts, size=8, weight="bold")
                ax.set_title("Modules Vs Bugs Count")
                if Bugs_CountListSum == 0:
                    print("Bugs_CountListSumCount is 0")
                else:
                    #plt.gcf().set_size_inches(11, 10)
                    plt.savefig(UserKeys[user]+'_ModuleVsBugsCount.jpg', dpi=500)
        else:
            print("Graph_show is " +Graph_show)
        Column_Name.clear()
        ModuleName.clear()

        Bugs_Count.clear()
        Bugs_CountList.clear()
        ColumnNumber = 0
        data.clear()
        plt.close()
        print("Cleared data " + str(user) + " for user " + UserKeys[user])

    except Exception as pp:
        print("Report File not found for "+UserKeys[user])
        print(pp)

        ExcelFileName = "UserData"
        loc = (ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        Sheetname = "General"
        sheetx = wb[Sheetname]
        for ix in range(1, 200):
            if sheetx.cell(ix, 1).value == None:
                break
            else:
                try:
                    if sheetx.cell(ix, 1).value == "Email_Content_NoFile":
                        print("Email_Content_NoFile is: " + sheetx.cell(ix, 2).value)
                        Email_Content_NoFile = sheetx.cell(ix, 2).value
                    if sheetx.cell(ix, 1).value == "email_from_NoFile":
                        print("email_from_NoFile is: " + sheetx.cell(ix, 2).value)
                        email_from_NoFile = sheetx.cell(ix, 2).value
                    if sheetx.cell(ix, 1).value == "Email_From_NoFile":
                        print("Email_From_NoFile is: " + sheetx.cell(ix, 2).value)
                        Email_From_NoFile = sheetx.cell(ix, 2).value
                    if sheetx.cell(ix, 1).value == "Email_Subject_NoFile":
                        print("Email_Subject_NoFile is: " + sheetx.cell(ix, 2).value)
                        Email_Subject_NoFile = sheetx.cell(ix, 2).value
                    if sheetx.cell(ix, 1).value == "GoogleAppCode_NoFile":
                        print("GoogleAppCode_NoFile is: " + sheetx.cell(ix, 2).value)
                        GoogleAppCode_NoFile = sheetx.cell(ix, 2).value
                    if sheetx.cell(ix, 1).value == "CC_Email":
                        print("CC_Email is: " + sheetx.cell(ix, 2).value)
                        CC_Email = sheetx.cell(ix, 2).value
                except Exception as ad:
                    print("No Data found " + str(ad))
        try:
            html = '''
                        <html>
                            <body>
                                <p>Hi ''' + UserKeys[user] + '''</p 
                                <p>''' + Email_Content_NoFile + '''<br /><br /></p>
                                <p>Many Thanks <br/>BIG QA Manager</p>
                            </body>
                        </html>
                        '''

            y = User_Name_Email[UserKeys[user]]
            SenderEmail = Email_From_NoFile
            date_str = pd.Timestamp.today().strftime('%m-%d-%Y')
            msg = MIMEMultipart()
            rcpt = [y] + [CC_Email]
            msg['Subject'] = UserKeys[user]+ "-"+Email_Subject_NoFile + " " + date_str
            msg['From'] = email_from_NoFile
            msg['Cc'] = CC_Email
            msg['To'] = y

            msg.attach(MIMEText(html, "html"))

            # ------------------------To attach all in e-Mail-----------------------
            email_string = msg.as_string()
            # -----------------------------------------------------------------------

            # ----------------------------SMTP setup--------------------------------
            server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
            RandmStr = GoogleAppCode_NoFile
            server.login(SenderEmail, RandmStr)
            server.sendmail(email_from_NoFile, rcpt, email_string)
            print("No Report email sent for " + UserKeys[user])
            server.quit()

        except Exception as aq:
            print("Email not sent for " + UserKeys[user])
            print(aq)

