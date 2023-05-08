import time
import glob
from pathlib import Path
import os
import openpyxl

home = str(Path.home())
home = home.replace(os.sep, '/')
print(home)

ExcelFileName = "UserData"
loc = (ExcelFileName + '.xlsx')
wb = openpyxl.load_workbook(loc)

Sheetname="Data"
sheet = wb[Sheetname]
User_Name_Sheet={}
User_Name_Email={}
for x in range(2, 200):
    if sheet.cell(x,1).value == None:
        break
    else:
        User_Name_Sheet[sheet.cell(x, 1).value]=sheet.cell(x, 2).value
        User_Name_Email[sheet.cell(x, 1).value]=sheet.cell(x, 3).value

print(User_Name_Sheet)
UserKeys=list(User_Name_Sheet.keys())
print(UserKeys)
for user in range(0,len(UserKeys)):
    try:
        os.remove(home+'/.jenkins/workspace/Create_Graph/'+UserKeys[user]+'_ModuleVsBugsCount.jpg')
    except:
        print("ModuleVsBugsCount not able to delete for "+UserKeys[user])
#-----------------To delete pdf and report xlsx files----------------------------
time.sleep(2)
ii=0
fileList1 = glob.glob(home+'/.jenkins/workspace/CreateReport/*.pdf')
fileList2 = glob.glob("ReportData/"+'/*.xlsx')

for ii in range(0,len(fileList1)):
    try:
        os.remove(fileList1[ii])
        os.remove(fileList2[ii])
    except Exception as ae:
        print(ae)
        print("No Attachment found to delete")

# Deleting Report files from Git Report folder
from github import Github

ExcelFileName = "GitAccessToken"
loc = (home+"/"+ExcelFileName + '.xlsx')
wb = openpyxl.load_workbook(loc)
Sheetname="Cred"
sheetx = wb[Sheetname]
for ix in range(1, 200):
    if sheetx.cell(ix, 1).value == None:
        break
    else:
        if sheetx.cell(ix, 1).value == "Git_Token":
            print("Git_Token is: "+sheetx.cell(ix, 2).value)
            Git_Token=sheetx.cell(ix, 2).value
        if sheetx.cell(ix, 1).value == "Git_Username":
            print("Git_Username is: "+sheetx.cell(ix, 2).value)
            Git_Username=sheetx.cell(ix, 2).value
        if sheetx.cell(ix, 1).value == "Git_Password":
            print("Git_Password is: "+sheetx.cell(ix, 2).value)
            Git_Password=sheetx.cell(ix, 2).value

g = Github(Git_Username, Git_Password)
g = Github(Git_Token)
#  Get all repos present
for repo in g.get_user().get_repos():
    print(repo.name)

# Accessing particular Repo and its folder
repo=g.get_repo("Neeraj5690/Reporting")
# Removing files from the folder
Folder=repo.get_contents("/ReportData")
for contentFiles in Folder:
    print(contentFiles)
    if contentFiles.path.format() == "ReportData/img.png":
        pass
    else:
        repo.delete_file(contentFiles.path, "message", contentFiles.sha, branch='master')
