from builtins import print
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import pandas as pd
import smtplib

Email_Content="Good Evening !!!  Please don't forget to add report file for this week."
FileLink="document/d/1n6lD3vMiOFaUved1Q_a6QsBoKKjFaKZJ8Ye8hrJL2z0/edit?usp=share_link"
email_from = 'Test Automation Team'
Email_From="neeraj1wayitsol@gmail.com"
Email_Subject="Weekly Report Reminder Email"
GoogleAppCode="tsiajyfnhywxctwi"

ExcelFileName = "UserData"
loc = (ExcelFileName + '.xlsx')
wb = openpyxl.load_workbook(loc)

Sheetname="General"
sheetx = wb[Sheetname]
for ix in range(1, 200):
    if sheetx.cell(ix, 1).value == None:
        break
    else:
        try:
            if sheetx.cell(ix, 1).value == "Email_Content":
                print("Email_Content is: "+sheetx.cell(ix, 2).value)
                Email_Content=sheetx.cell(ix, 2).value
            if sheetx.cell(ix, 1).value == "FileLink":
                print("FileLink is: "+sheetx.cell(ix, 2).value)
                FileLink=sheetx.cell(ix, 2).value
            if sheetx.cell(ix, 1).value == "email_from":
                print("email_from is: "+sheetx.cell(ix, 2).value)
                email_from=sheetx.cell(ix, 2).value
            if sheetx.cell(ix, 1).value == "Email_From":
                print("Email_From is: "+sheetx.cell(ix, 2).value)
                Email_From=sheetx.cell(ix, 2).value
            if sheetx.cell(ix, 1).value == "Email_Subject":
                print("Email_Subject is: "+sheetx.cell(ix, 2).value)
                Email_Subject=sheetx.cell(ix, 2).value
            if sheetx.cell(ix, 1).value == "GoogleAppCode":
                print("GoogleAppCode is: "+sheetx.cell(ix, 2).value)
                GoogleAppCode=sheetx.cell(ix, 2).value
            if sheetx.cell(ix, 1).value == "NewVersion":
                print("NewVersion is: " + sheetx.cell(ix, 2).value)
                NewVersion = sheetx.cell(ix, 2).value
                NewVersionValue = sheetx.cell(ix, 3).value
        except Exception as ad:
            print("No Data found "+str(ad))

Sheetname="Data"
sheet = wb[Sheetname]
User_Name_Sheet={}
User_Name_Email={}
for x in range(2, 200):
    if sheet.cell(x,1).value == None:
        break
    else:
        User_Name_Sheet[sheet.cell(x, 1).value]=sheet.cell(x, 2).value
        User_Name_Email[sheet.cell(x, 1).value]=sheet.cell(x, 3).value

#print(User_Name_Sheet)
UserKeys=list(User_Name_Sheet.keys())
#print(UserKeys)
for user in range(0,len(UserKeys)):
    try:
        if NewVersion == "No":
            html = '''
                        <html>
                            <body>
                                <p>Hi '''+UserKeys[user]+'''</p 
                                <p>Good Evening !!!  </p
                                <p>''' + Email_Content + '''<br /></p>
                                <p>To know more how to add file to GitHub folder checkout the link given below <br /></p>
                                <p>''' + "https://docs.google.com/" + FileLink + '''<br /><br /></p>
                                <p>Many Thanks <br/>Neeraj</p>
                            </body>
                        </html>
                        '''
        elif NewVersion == "Yes":
            html = '''
                                    <html>
                                        <body>
                                            <p>Hi ''' + UserKeys[user] + '''</p 
                                            <p>Good Evening !!!  </p
                                            <p>New version released. To check more please go to the link: ''' + NewVersionValue + '''<br /></p
                                            <p>''' + Email_Content + '''<br /></p>
                                            <p>To know more how to add file to GitHub folder checkout the link given below <br /></p>
                                            <p>''' + "https://docs.google.com/" + FileLink + '''<br /><br /></p>
                                            <p>Many Thanks <br/>Neeraj</p>
                                        </body>
                                    </html>
                                    '''
        y = User_Name_Email[UserKeys[user]]
        SenderEmail = Email_From
        date_str = pd.Timestamp.today().strftime('%m-%d-%Y')
        msg = MIMEMultipart()
        msg['Subject'] = Email_Subject + " " + date_str
        msg['From'] = email_from
        msg['To'] = y

        msg.attach(MIMEText(html, "html"))

        # ------------------------To attach all in e-Mail-----------------------
        email_string = msg.as_string()
        # -----------------------------------------------------------------------

        # ----------------------------SMTP setup--------------------------------
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        RandmStr = GoogleAppCode
        server.login(SenderEmail, RandmStr)
        server.sendmail(email_from, y, email_string)
        print("Reminder email sent for "+UserKeys[user])
        server.quit()

    except Exception as dd:
        print("Email not sent for "+UserKeys[user])
        print(dd)