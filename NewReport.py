import os
import openpyxl
from fpdf import FPDF, fpdf
import datetime
import pandas as pd
from pathlib import Path

home = str(Path.home())
home = home.replace(os.sep, '/')
print(home)

version = " version: 1.02 "

date_str = pd.Timestamp.today().strftime('%m-%d-%Y')
ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M %p")

ExcelFileName = "UserData"
loc = (ExcelFileName + '.xlsx')
wb = openpyxl.load_workbook(loc)

Sheetname="Data"
sheet = wb[Sheetname]
User_Name={}
for x in range(2, 200):
    if sheet.cell(x,1).value == None:
        break
    else:
        print("User_Name : "+sheet.cell(x, 1).value)
        print("User_File : "+sheet.cell(x, 2).value)
        User_Name[sheet.cell(x, 1).value]=sheet.cell(x, 2).value

print(User_Name)
UserKeys=list(User_Name.keys())
print(UserKeys)

for user in range(0,len(UserKeys)):
    # print(UserKeys[user])
    # print(User_Name[UserKeys[user]])
    Report_Title = "Report"
    Project_Name = "QA"
    Report_Name = "Report"

    try:
        ExcelFileName = "ReportData/"+User_Name[UserKeys[user]]

        # Connecting with Main Report Data File
        locx = (ExcelFileName + '.xlsx')
        wbx = openpyxl.load_workbook(locx)

        # Reading GlobalData tab of Main Report Data File
        Sheetname="GlobalData"
        sheetx = wbx[Sheetname]
        for ix in range(1, 200):
            if sheetx.cell(ix, 1).value == None:
                break
            else:
                if sheetx.cell(ix, 1).value == "Project_Name":
                    print("Project Name is: "+sheetx.cell(ix, 2).value)
                    Project_Name=sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "Report_Title":
                    print("Report Title is: "+sheetx.cell(ix, 2).value)
                    Report_Title=sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "Report_Name":
                    print("Report_Name is: "+sheetx.cell(ix, 2).value)
                    Report_Name=sheetx.cell(ix, 2).value+'_'+date_str+'.pdf'
                if sheetx.cell(ix, 1).value == "Project_Status":
                    print("Project_Status is: "+sheetx.cell(ix, 2).value)
                    Project_Status=sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "Graph_show":
                    print("Graph_show is: "+sheetx.cell(ix, 2).value)
                    Graph_show=sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "Graph_Type":
                    print("Graph_Type is: "+sheetx.cell(ix, 2).value)
                    Graph_Type=sheetx.cell(ix, 2).value
                if sheetx.cell(ix, 1).value == "BugCount_show":
                    print("BugCount_show is: "+sheetx.cell(ix, 2).value)
                    BugCount_show=sheetx.cell(ix, 2).value

        # Reading Column Name data
        Sheetname="ModulesData"
        sheetx = wbx[Sheetname]
        Column_Name = []
        for ix2 in range(1, 20):

            if sheetx.cell(1,ix2).value == None:
                break
            else:
                Column_Name.append(sheetx.cell(1,ix2).value)
        print("Column_Name "+str(Column_Name))

        # Reading Module Name data
        ModuleName=[]

        # Reading Bugs_Count data
        Bugs_Count={}
        Bugs_CountList=[]

        # Reading Bugs_Links data
        Bugs_Links={}

        # Reading Triage data
        Comment={}

        for ix1 in range(1, len(Column_Name)+1):
            if ix1==1:
                LastCell=100
            else:
                LastCell = len(ModuleName)

            for ix11 in range(2,LastCell+2 ):
                # Do not change this - it is fetching all the module names
                if ix1==1:
                    if sheetx.cell(ix11, ix1).value == None:
                        break
                    else:
                        ModuleName.append(sheetx.cell(ix11,ix1).value)
                else:
                    # Add new parameter here if required
                    if ix1==2:
                        if ix11 == len(ModuleName)+2:
                            break
                        if sheetx.cell(ix11,ix1).value == None:
                            #print("sheetx.cell(ix11,ix1).value -> "+sheetx.cell(ix11,ix1).value)
                            Bugs_Count[ModuleName[ix11 - 2]] = -1
                            Bugs_CountList.append(0)
                        else:
                            Bugs_Count[ModuleName[ix11 - 2]] = sheetx.cell(ix11,ix1).value
                            Bugs_CountList.append(sheetx.cell(ix11,ix1).value)
                    if ix1==3:
                        if ix11 == len(ModuleName)+2:
                            break
                        if sheetx.cell(ix11,ix1).value == None:
                            Bugs_Links[ModuleName[ix11 - 2]] = "None"
                        else:
                            Bugs_Links[ModuleName[ix11 - 2]] = sheetx.cell(ix11,ix1).value
                    if ix1==4:
                        if ix11 == len(ModuleName)+2:
                            break
                        if sheetx.cell(ix11,ix1).value == None:
                            Comment[ModuleName[ix11 - 2]] = "None"
                        else:
                            Comment[ModuleName[ix11 - 2]] = sheetx.cell(ix11,ix1).value

        print("Modules are "+str(ModuleName))
        print("Bugs_Count "+str(Bugs_Count))
        print("Bugs_CountList " + str(Bugs_CountList))
        print("Sum of Bugs_CountList " + str(sum(Bugs_CountList)))
        MaxBugs = max(Bugs_Count, key=Bugs_Count.get)
        print("MaxBugs "+str(MaxBugs))
        print("Bugs_Links are "+str(Bugs_Links))
        print("Comments are "+str(Comment))

        class PDF(FPDF):
            def header(self):
                self.image('Logo.png', 10, 8, 33)
                self.add_font('Arial', '', 'C:\Windows\Fonts\Arial.ttf', uni=True)
                self.set_font('Arial', 'B', 15)
                self.set_y(16)
                w = self.get_string_width(Report_Title+": "+Project_Name) + 6
                self.set_x((210 - w) / 2)
                self.set_draw_color(0, 80, 180)
                self.set_fill_color(0,76,153)
                self.set_text_color(255,255,255)
                self.set_line_width(1)
                self.cell(w, 7, Report_Title+": "+Project_Name, 1, 1, 'C', 1)
                self.ln(5)

            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.set_text_color(128)
                self.cell(0, 10, 'Page ' + str(self.page_no()), 0, 0, 'C')

            def section_counts(self):
                MaxBugs="N/A"
                if BugCount_show == "No":
                    SumBugCount = "N/A"
                    MaxBugs = "N/A"
                else:
                    if sum(Bugs_CountList) == 0:
                        SumBugCount = "N/A"
                        MaxBugs = "N/A"
                    else:
                        SumBugCount = str(sum(Bugs_CountList))
                        MaxBugs = max(Bugs_Count, key=Bugs_Count.get)

                self.ln(17)
                self.set_font("Arial", size=12)
                w1 = self.get_string_width(" Total Bugs : " + SumBugCount) + 3
                X = (70 - w1) / 2
                self.set_x(X)

                Y = 36
                multplyVar=15
                self.set_y(Y)
                self.set_draw_color(0, 80, 180)
                self.set_fill_color(198, 224, 228)
                self.set_text_color(0, 0, 0)
                self.set_line_width(1)
                self.cell(w1, 9, " Total bugs : "+SumBugCount, 1, 1, 'L', 1)

                Y=Y+multplyVar
                self.set_y(Y)
                self.set_draw_color(0, 80, 180)
                self.set_fill_color(231, 203, 149)
                self.set_text_color(0, 0, 0)
                self.set_line_width(1)
                w2 = self.get_string_width(" Max bugs from : "+MaxBugs) + 3
                self.cell(w2, 9, " Max bugs from : "+MaxBugs, 1, 1, 'L', 1)

                if Graph_show=="Yes":
                    Y = Y + multplyVar+17
                    self.set_y(Y)
                    X = 10
                elif Graph_show=="No":
                    X = 120
                    Y=35
                    self.set_x(X)
                    self.set_y(Y)
                self.set_draw_color(0, 80, 180)
                self.set_fill_color(0, 76, 153)
                self.set_text_color(255, 255, 255)
                self.set_x(X)
                w2 = self.get_string_width(" Overall Status ") + 3
                self.cell(w2, 9, " Overall Status ", 1, 0, 'L', 1)


                OverallStatusText = "None"

                self.set_fill_color(224, 224, 224)
                Y = Y + 11
                self.set_y(Y)
                self.set_x(X)
                self.set_draw_color(0, 80, 180)
                if Project_Status   == "Off Track":
                    self.set_fill_color(255, 51, 51)
                    OverallStatusText = Project_Status
                self.set_text_color(0, 0, 0)
                self.cell(9, 9, " ", 0, 0, 'L', 1)

                self.set_fill_color(224, 224, 224)
                X = X + 11
                self.set_x(X)
                if Project_Status == "Concerned":
                    self.set_fill_color(255, 153, 51)
                    OverallStatusText = Project_Status
                self.set_text_color(0, 0, 0)
                self.cell(9, 9, " ", 0, 0, 'L', 1)

                self.set_fill_color(224, 224, 224)
                X = X + 11
                self.set_x(X)
                if Project_Status == "On Track":
                    self.set_fill_color(0, 204, 0)
                    OverallStatusText = Project_Status
                self.set_text_color(0, 0, 0)
                self.cell(9, 9, " ", 0, 0, 'L', 1)

                self.set_x(10)
                Y = Y + 12
                self.set_y(Y)
                self.set_x(X-22)
                self.set_font("Arial", size=8)
                self.set_text_color(80, 90, 80)
                self.cell(0, 0,"( "+OverallStatusText+" ) ", 0, 0, 'L')

                self.ln(15)
                print("Y is "+str(Y))
                if Graph_show=="Yes":
                    if Y <= 120:
                        Y = 120
                    else:
                        Y = Y + multplyVar + 15
                elif Graph_show=="No":
                    Y = 80

                self.set_y(Y)
                self.set_font('Arial', 'B', 12)
                self.set_text_color(0, 0, 0)
                self.set_fill_color(200, 220, 255)
                self.cell(0, 0, 'Module details : ', 0, fill=False)
                self.ln(5)

            def section_title(self, num, label):
                self.set_font('Arial', '', 12)
                self.set_text_color(0, 0, 0)
                self.set_fill_color(200, 220, 255)
                self.cell(0, 10, 'Module %d : %s' % (num, label), 0, 1, 'L', 1)

                if BugCount_show == "No":
                    pass
                elif BugCount_show == "Yes":
                    if str(Bugs_Count[label]) == "-1":
                        pass
                        #self.cell(0, 10, 'Bugs Count : ', 0, 1)
                    else:
                        self.cell(0, 10, 'Bugs Count : ' + str(Bugs_Count[label]), 0, 1)

                self.set_text_color(0, 0, 255)
                if str(Bugs_Links[label]) == "None":
                    pass
                else:
                    self.cell(0, 6, 'Link : ' + str(Bugs_Links[label]), 0, 1)
                self.set_text_color(0, 0, 0)

                self.cell(0, 7, 'Details : ', 0, 1)
                self.set_font('Arial', '', 9)
                if str(Comment[label]) == "None":
                    pass
                else:
                    self.multi_cell(0, 5, str(Comment[label]), 0)
                self.ln(5)

            def Graph(self):
                self.set_font("Arial", size=8)
                self.set_text_color(80, 90, 80)
                self.cell(0, 1, version +"    Report Date: "+ctReportHeader, 0, 0, 'L')
                try:
                    if Graph_Type == "BarGraph":
                        self.image(home+'/.jenkins/workspace/Create_Graph/'+UserKeys[user]+'_ModuleVsBugsCount.jpg', 100, 25, 100,90)
                    elif Graph_Type == "PieChart":
                        self.image(home+'/.jenkins/workspace/Create_Graph/'+UserKeys[user]+'_ModuleVsBugsCount.jpg', 80, 40, 120,60)
                except Exception as aa:
                    print("aa "+str(aa))
                    try:
                        if Graph_Type == "BarGraph":
                            self.image(UserKeys[user]+'_ModuleVsBugsCount.jpg', 100, 25, 100,90)
                        elif Graph_Type == "PieChart":
                            self.image(UserKeys[user]+'_ModuleVsBugsCount.jpg', 80, 40, 120,60)
                    except Exception as av:
                        print("av "+str(av))
                        if Graph_show == "Yes":
                            try:
                                self.image(home + '/.jenkins/workspace/Create_Graph/' +'NoData.jpg', 100, 25, 100, 90)
                            except :
                                self.image('NoData.jpg', 100, 25, 100, 90)

            def print_Data(self, num, Report_Title):
                self.section_title(num, Report_Title)

        pdf = PDF()
        pdf.set_title(Report_Title)
        pdf.add_page()
        pdf.Graph()
        pdf.section_counts()

        # Module Data
        Sheetname="ModulesData"
        sheetx = wbx[Sheetname]
        for ix in range(1, len(ModuleName)+1):
            pdf.print_Data(ix, ModuleName[ix-1])

        pdf.output(UserKeys[user]+"_"+Report_Name)
        Column_Name.clear()
        ModuleName.clear()
        Bugs_Count.clear()
        Bugs_CountList.clear()
        Bugs_Links.clear()
        Comment.clear()
        Project_Status = None
        MaxBugs = None

    except Exception as aaa:
        print("Report File not found for "+UserKeys[user])
        print(aaa)

